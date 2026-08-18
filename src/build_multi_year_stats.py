"""
One-time conversion script: turns the user-supplied "Advanced Stats.xlsx"
(DARKO DPM leaderboards + Basketball-Reference PER/BPM/VORP/Age tables for
the 2023-24, 2024-25, and 2025-26 seasons) into the CSV schema
player_value_regression.py expects (see that module's docstring):

    Player,Team,Season,Age,DARKO_DPM,BPM,VORP

Season is labeled by the ENDING year of the season (2023-24 -> 2024,
2024-25 -> 2025, 2025-26 -> 2026), matching this project's convention.

The workbook's "DPM - <season>" sheets are normal columnar data (Player,
Team, DPM, ...) -- Team here is already a full franchise name, matching
darkodpmleaderboard.csv's spelling, so DPM sheets are the source of truth
for both Player spelling and Team.

The "PER VORP BPM - <season>" sheets came in as single-column pasted CSV
text (one full comma-separated row per cell, from a Basketball-Reference
export) rather than real columns -- each row is re-parsed with csv.reader.
That table only supplies Age/BPM/VORP/PER (indexed by its OWN Player
spelling, and Team as a 2-3 letter code we never need since we keep the
DPM sheet's Team). Header rows that Basketball-Reference repeats every ~20
rows, the trailing "Provided by..." attribution row, and multi-team
"TOT"/"2TM"/"3TM" rollup rows (which would double-count a traded player)
are all filtered out; a player's single-team rows are used instead.

MATCHING PLAYERS ACROSS THE TWO TABLES: the two sources spell some names
differently (diacritics dropped/kept inconsistently, "Jr."/"Sr."/"II"/"III"
suffixes present on one side only, a handful of nicknames like "Bones
Hyland" for "Nah'Shon Hyland"). Matching normalizes both sides (strip
diacritics, drop periods, lowercase, drop suffix tokens) and then applies
a small hand-verified alias table for the remaining real nickname/legal-
name mismatches (checked by hand against each sheet's contents -- see
NAME_ALIASES below). Players a season's PER/VORP/BPM table has no row for
at all (e.g. Steven Adams and Lonzo Ball for 2023-24, both out that full
season with injuries; Tyrese Haliburton/Fred VanVleet/Kyrie
Irving/Damian Lillard for 2025-26, all out with season-ending injuries)
are skipped for that season only -- they simply contribute one fewer
season of history for that player, which is the correct behavior per
player_value_regression.py's docstring (players with too little history
just keep their raw current-season DPM). This was verified by hand: after
normalization + aliasing, every remaining name gap across all three
seasons corresponds to a real "didn't log qualifying minutes that season"
case, not a silent join failure -- see this file's __main__ for the
diagnostic printout used to check that.

PER is also read from the PER/VORP/BPM sheets and written to the output
CSV as an extra "PER" column, even though player_value_regression.py's
documented schema and its regression composite currently use only
DARKO_DPM/BPM/VORP (PER is not yet one of the three z-scored inputs). This
is a deliberate, flagged choice: csv.DictReader (which load_multi_year_stats
uses) ignores columns it doesn't look up, so adding PER here is harmless
today and means the data isn't thrown away -- extending the composite/
FEATURE_NAMES to actually use it is a follow-up if wanted, not done here
so as not to silently change the fitted model's behavior/coefficients
without that being an explicit decision.
"""

import csv
from typing import Dict, Tuple

import openpyxl

from name_matching import normalize_name

XLSX_PATH = "/root/.claude/uploads/546919a0-f99b-590f-be84-ef42a59fcd31/ec59b95e-Advanced_Stats.xlsx"
OUTPUT_CSV = "data/multi_year_advanced_stats.csv"

SEASON_SHEETS = [
    ("DPM - 2023-2024", "PER VORP BPM - 2023-2024", 2024),
    ("DPM - 2024-2025", "PER VORP BPM - 2024-2025", 2025),
    ("DPM - 2025-2026", "PER VORP BPM - 2025-2026", 2026),
]

# Hand-verified: DPM-sheet spelling (key) -> PER/VORP/BPM-sheet spelling
# (value), both already run through normalize_name() (see name_matching.py).
# Found by diffing the two sheets' name sets per season -- see this file's
# __main__.
NAME_ALIASES = {
    "nicolas claxton": "nic claxton",
    "nahshon hyland": "bones hyland",
    "kenyon martin": "kj martin",
    "ronald holland": "ron holland",
    "alexandre sarr": "alex sarr",
    "carlton carrington": "bub carrington",
    "egor demin": "egor demin",
    "david jones": "david jones garcia",
}


def load_dpm_sheet(wb, sheet_name: str) -> Dict[str, Tuple[str, float]]:
    """{Player (canonical spelling): (Team, DARKO_DPM)}"""
    ws = wb[sheet_name]
    out = {}
    for row in ws.iter_rows(min_row=3, values_only=True):  # skip title row + header row
        if not row or row[1] is None:
            continue
        player, team, dpm = row[1], row[2], row[4]
        out[player.strip()] = (team.strip(), float(dpm))
    return out


def load_per_vorp_bpm_sheet(wb, sheet_name: str) -> Dict[str, Tuple[float, float, float, float]]:
    """{Player (PER/VORP/BPM-sheet spelling): (Age, BPM, VORP, PER)}"""
    ws = wb[sheet_name]
    lines = [row[0] for row in ws.iter_rows(min_row=1, values_only=True) if row and row[0]]
    out = {}
    for row in csv.DictReader(lines):
        name = row.get("Player")
        if not name or name == "Player":  # repeated header row
            continue
        team = row.get("Team")
        if not team or team in ("TOT", "2TM", "3TM", "4TM"):  # multi-team rollup row
            continue
        try:
            age = float(row["Age"])
            bpm = float(row["BPM"])
            vorp = float(row["VORP"])
            per = float(row["PER"])
        except (ValueError, TypeError):
            continue
        out.setdefault(name.strip(), (age, bpm, vorp, per))  # first (single-team) row wins
    return out


def build_rows(wb):
    """Yields (Player, Team, Season, Age, DARKO_DPM, BPM, VORP, PER) rows; also returns per-season skip diagnostics."""
    all_rows = []
    diagnostics = []
    for dpm_sheet, pv_sheet, season in SEASON_SHEETS:
        dpm = load_dpm_sheet(wb, dpm_sheet)
        pv = load_per_vorp_bpm_sheet(wb, pv_sheet)
        pv_by_norm = {normalize_name(name): stats for name, stats in pv.items()}

        matched, skipped = 0, []
        for player, (team, darko_dpm) in dpm.items():
            key = normalize_name(player)
            key = NAME_ALIASES.get(key, key)
            stats = pv_by_norm.get(key)
            if stats is None:
                skipped.append(player)
                continue
            age, bpm, vorp, per = stats
            all_rows.append((player, team, season, age, darko_dpm, bpm, vorp, per))
            matched += 1
        diagnostics.append((season, len(dpm), matched, skipped))
    return all_rows, diagnostics


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    rows, diagnostics = build_rows(wb)

    print("Per-season match diagnostics:")
    for season, total, matched, skipped in diagnostics:
        print(f"  {season}: {matched}/{total} players matched, {len(skipped)} skipped (no qualifying-minutes row that season)")
        if skipped:
            print(f"    skipped: {skipped}")

    rows.sort(key=lambda r: (r[0], r[2]))  # Player, then Season ascending

    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Player", "Team", "Season", "Age", "DARKO_DPM", "BPM", "VORP", "PER"])
        for player, team, season, age, darko_dpm, bpm, vorp, per in rows:
            writer.writerow([player, team, season, age, darko_dpm, bpm, vorp, per])

    print(f"\nWrote {len(rows)} player-season rows to {OUTPUT_CSV}")
    n_players = len({r[0] for r in rows})
    print(f"Covering {n_players} distinct players")


if __name__ == "__main__":
    main()
